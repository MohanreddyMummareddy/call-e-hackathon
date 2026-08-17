"""CALL-E Hackathon: Customer Follow-Up Automation

Python SDK + Google Calendar API integration for automated lead follow-up calls.
Reference demo: preview mode is fully simulated; live mode requires explicit
operator opt-in and per-call confirmation. See README for operational notes.
"""

import os
import sys
import json
import csv
import io
import time
import threading
import asyncio
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask import Flask, request, jsonify, render_template_string, session, redirect, send_file
from dotenv import load_dotenv
from call_e_sdk import CALL_EClient
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request

load_dotenv()

# Flask app for webhook endpoints
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "dev-secret-key-for-hackathon")

# CONFIGURATION
# No placeholder: an empty/missing key must fail closed (never look configured).
CALL_E_API_KEY = os.environ.get("CALL_E_API_KEY", "")
if CALL_E_API_KEY in ("your-api-key-here", "YOUR_API_KEY_HERE"):
    CALL_E_API_KEY = ""
call_e_client = CALL_EClient(api_key=CALL_E_API_KEY)

# Explicit server-side live enable. Credential presence alone never selects
# live behavior: calls are placed only when CALLE_LIVE_CALLS_ENABLED is true.
LIVE_CALLS_ENABLED = os.environ.get("CALLE_LIVE_CALLS_ENABLED", "").strip().lower() in ("1", "true", "yes")

# Bearer token required by every state-changing/private route. The demo UI
# collects it once and sends it as "Authorization: Bearer <token>".
APP_TOKEN = os.environ.get("APP_TOKEN", "").strip()
if LIVE_CALLS_ENABLED and (len(APP_TOKEN) < 16 or any(ch.isspace() for ch in APP_TOKEN)):
    raise RuntimeError(
        "CALLE_LIVE_CALLS_ENABLED requires APP_TOKEN: a whitespace-free secret of at least 16 characters."
    )

# Explicit operator opt-in for persisting the OAuth token to disk / restoring
# from GOOGLE_TOKEN_JSON. Without this, tokens live only in server memory
# keyed by the session that authorized them.
PERSIST_TOKEN = os.environ.get("PERSIST_GOOGLE_TOKEN", "").strip().lower() in ("1", "true", "yes")

# Google Calendar Setup - OAuth 2.0
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/calendar.events",
    "https://www.googleapis.com/auth/gmail.send",
]
GOOGLE_OAUTH_CONFIG = {
    "web": {
        "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
        "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET"),
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
        "redirect_uris": [
            os.environ.get("GOOGLE_REDIRECT_URI_LOCAL", "http://localhost:8080/oauth2callback"),
        ],
    }
}
# Choose which registered redirect URI the OAuth flow uses (local first, prod override)
GOOGLE_REDIRECT_URI = os.environ.get("GOOGLE_REDIRECT_URI", GOOGLE_OAUTH_CONFIG["web"]["redirect_uris"][0])

# Server-side OAuth token store keyed by the session that authorized it.
# Persistence to disk happens only when PERSIST_GOOGLE_TOKEN=true (operator opt-in).
CREDENTIALS_FILE = "user_token.json"
oauth_tokens = {}

# When present, suppress GOOGLE_TOKEN_JSON restore so Disconnect sticks
DISCONNECT_FLAG = "oauth_disconnected.flag"

# In-memory pending lead store keyed by call SID (single-process demo)
pending_leads = {}

# Batch upload jobs keyed by batch id
batch_jobs = {}

# ============================================================
# AUTH / PRIVACY HELPERS
# ============================================================

def require_token(strict=False):
    """Fail closed on every protected route: a valid Bearer APP_TOKEN is mandatory.

    `strict=True` (call-result, OAuth, batch, and data routes) always requires
    authentication, even in the credential-free preview: an operator may run a
    preview with a provider key present, and an unauthenticated caller must not
    be able to query call status/results or exported data in that configuration.

    The non-strict gate (lead submission only) opens when live calls are
    disabled AND the operator has not set APP_TOKEN, so the documented fake
    preview stays runnable without any credentials. Nothing can be dialed,
    booked, or persisted in that mode.
    """
    if not strict and not LIVE_CALLS_ENABLED and not APP_TOKEN:
        return None
    auth = request.headers.get("Authorization", "")
    expected = "Bearer " + APP_TOKEN
    if not APP_TOKEN or auth != expected:
        return jsonify({"error": "Authentication required. Send 'Authorization: Bearer <APP_TOKEN>'."}), 401
    return None


def current_session_id():
    """Stable per-browser session id stored in the signed Flask session cookie."""
    sid = session.get("_sid")
    if not sid:
        sid = os.urandom(16).hex()
        session["_sid"] = sid
    return sid


def validate_e164(phone):
    """Return the phone as strict E.164, or None when not a valid real number."""
    try:
        import phonenumbers
        num = phonenumbers.parse(phone or "", None)
        if phonenumbers.is_possible_number(num) and phonenumbers.is_valid_number(num):
            return phonenumbers.format_number(num, phonenumbers.PhoneNumberFormat.E164)
    except Exception:
        pass
    return None


def mask_phone(phone):
    """Mask a phone number for logs/responses: +91 •••• •••• 3210."""
    if not phone:
        return ""
    try:
        import phonenumbers as _pn
        num = _pn.parse(phone.replace(" ", ""), None)
        cc = "+" + str(num.country_code)
        digits = str(num.national_number)
    except Exception:
        import re
        m = re.match(r"^(\+\d{1,3})(\d+)$", phone.replace(" ", ""))
        if not m:
            return phone
        cc, digits = m.group(1), m.group(2)
    if len(digits) <= 4:
        return cc + "••••"
    return cc + "•••• •••• " + digits[-4:]

# HTML template for interactive demo page
DEMO_HTML = """
<!DOCTYPE html>
<html>
<head>
    <title>CALL-E Customer Follow-Up Automation</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #f0f4f8; color: #1a202c; padding: 20px; }
        .container { max-width: 900px; margin: 0 auto; }
        h1 { font-size: 28px; margin-bottom: 4px; }
        .subtitle { color: #4a5568; margin-bottom: 24px; }
        .card { background: white; border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
        .card h2 { font-size: 18px; margin-bottom: 16px; color: #2d3748; }
        label { display: block; font-size: 14px; font-weight: 600; margin: 12px 0 4px; color: #4a5568; }
        input { width: 100%; padding: 10px 12px; border: 2px solid #e2e8f0; border-radius: 8px; font-size: 15px; transition: border 0.2s; }
        input:focus { outline: none; border-color: #4a90d9; }
        .btn { padding: 12px 20px; border: none; border-radius: 8px; font-size: 15px; font-weight: 600; cursor: pointer; margin-top: 16px; transition: opacity 0.2s; }
        .btn:hover { opacity: 0.85; }
        .btn-primary { background: #4a90d9; color: white; }
        .btn-green { background: #38a169; color: white; }
        .btn-gray { background: #a0aec0; color: white; }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; }
        .mode-btn { padding: 8px 16px; font-size: 14px; border-radius: 8px; border: none; cursor: pointer; }
        .mode-btn.active { background: #4a90d9 !important; color: white !important; }
        .status { margin-top: 16px; padding: 12px; border-radius: 8px; font-size: 14px; display: none; }
        .status.success { background: #f0fff4; border: 1px solid #38a169; color: #22543d; }
        .status.error { background: #fff5f5; border: 1px solid #e53e3e; color: #742a2a; }
        .status.info { background: #ebf8ff; border: 1px solid #4299e1; color: #2c5282; }
        .result-box { margin-top: 12px; padding: 12px; background: #f7fafc; border: 1px solid #e2e8f0; border-radius: 8px; font-family: monospace; font-size: 13px; white-space: pre-wrap; display: none; }
        .flow-step { display: flex; align-items: center; margin: 8px 0; padding: 10px; border-radius: 8px; background: #f7fafc; border: 1px solid #e2e8f0; }
        .flow-step .num { width: 28px; height: 28px; border-radius: 50%; background: #4a90d9; color: white; display: flex; align-items: center; justify-content: center; font-weight: 700; margin-right: 12px; flex-shrink: 0; }
        .flow-step.active { border-color: #4a90d9; background: #ebf8ff; }
        .flow-step.done { border-color: #38a169; background: #f0fff4; }
        .flow-step.done .num { background: #38a169; }
        .calendar-event { border: 2px solid #38a169; border-radius: 8px; padding: 12px; margin-top: 12px; background: #f0fff4; }
        .calendar-event h4 { color: #22543d; margin-bottom: 8px; }
        .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
        @media (max-width: 700px) { .grid-2 { grid-template-columns: 1fr; } }
    </style>
</head>
<body>
    <div class="container">
        <h1>CALL-E Customer Follow-Up Automation</h1>
        <p class="subtitle">Hybrid: CALL-E Python SDK + Google Calendar API</p>

        <div class="grid-2">
            <!-- LEFT: Lead submission form (toggle single vs bulk) -->
            <div class="card">
                <h2>1. New Lead / Call</h2>
                <div style="display:flex; gap:8px; margin-bottom:16px;">
                    <button class="btn mode-btn" id="mode-single" onclick="setMode('single')" style="margin-top:0; background:#4a90d9; color:white;">Single Lead</button>
                    <button class="btn mode-btn" id="mode-bulk" onclick="setMode('bulk')" style="margin-top:0; background:#a0aec0; color:white;">Bulk Upload</button>
                </div>
                <div class="status info" id="gate-msg" style="display:block;">Please connect Google Calendar &amp; Gmail first (right side) before placing calls.</div>

                <div id="single-mode">
                    <label>Your Company (caller)</label>
                    <input type="text" id="your-company" placeholder="Acme Corp" value="Acme Corp">
                    <label>Lead's Full Name</label>
                    <input type="text" id="name" placeholder="John Doe" value="John Doe">
                    <label>Lead's Phone Number</label>
                    <div style="display:flex; gap:8px;">
                        <select id="phone-country" style="width:38%; padding:10px 12px; border:2px solid #e2e8f0; border-radius:8px; font-size:15px; background:#fff;"></select>
                        <input type="text" id="phone" placeholder="55550100" style="flex:1; min-width:0; padding:10px 12px; border:2px solid #e2e8f0; border-radius:8px; font-size:15px;">
                    </div>
                    <div id="lead-tz-display" style="font-size:13px; color:#4a5568; margin-top:6px; min-height:18px;">Detected timezone: checking...</div>
                    <input type="hidden" id="company-tz" value="UTC">
                    <label>Lead's Email</label>
                    <input type="email" id="email" placeholder="john@example.com" value="john@example.com">
                    <label>Lead's Company (optional)</label>
                    <input type="text" id="company" placeholder="XYZ Inc" value="XYZ Inc">
                    <label>API Token (required)</label>
                    <input type="password" id="api-token" placeholder="APP_TOKEN set by the operator" autocomplete="off">
                    <label><input type="checkbox" id="live-confirm" style="width:auto; margin-right:6px;">I understand live calls are enabled server-side and this places a real phone call</label>
                    <label><input type="checkbox" id="lead-consent" style="width:auto; margin-right:6px;" checked>The lead requested this follow-up and consents to being called at this number</label>
                    <button class="btn btn-green" id="btn-call" onclick="submitLead()" disabled>Place CALL-E Follow-up Call</button>
                    <div class="status" id="call-status"></div>
                    <div class="result-box" id="call-result"></div>
                </div>

                <div id="bulk-mode" style="display:none;">
                    <p style="font-size:14px; color:#4a5568;">
                        Upload an Excel/CSV file with columns: <strong>name, phone, email, company, your_company</strong> (first row = headers, <strong>company_tz</strong> and <strong>consent</strong> optional; set consent=yes for every recipient you are authorized to call).
                        Calls happen sequentially. If a call's status becomes unknown (provider lookup failure or polling timeout), the batch <strong>stops</strong> and the remaining rows are marked <strong>stopped</strong>. After processing, download the file with a <strong>status</strong> column added.
                        Re-uploading the same file only calls leads that were <strong>not</strong> successfully scheduled.
                    </p>
                    <input type="file" id="batch-file" accept=".xlsx,.csv" style="padding:6px; border:1px solid #e2e8f0; border-radius:8px;">
                    <button class="btn btn-green" id="btn-batch" onclick="uploadBatch()">Upload &amp; Process Batch</button>
                    <button class="btn btn-gray" id="btn-batch-stop" onclick="stopBatch()" style="display:none;">Stop After Current Call</button>
                    <button class="btn btn-primary" id="btn-batch-download" onclick="downloadBatch()" style="display:none;">Download Updated Excel</button>
                    <div class="status info" id="batch-status" style="display:none;"></div>
                    <div class="result-box" id="batch-result" style="display:block;"></div>
                </div>
            </div>

            <!-- RIGHT: OAuth + Follow-up flow -->
            <div>
                <div class="card">
                    <h2>2. Google Calendar &amp; Gmail Authorization</h2>
                    <p style="font-size:14px; color:#4a5568;">Required before placing calls. Events are created on <strong>your</strong> calendar (company account) with the lead as attendee, and confirmation emails are sent via your Gmail.</p>
                    <div id="oauth-connected" style="display:none;">
                        <div class="status success" style="display:block;">Connected to: <span id="oauth-account"></span></div>
                        <p style="font-size:13px; color:#4a5568; margin-top:8px;" id="oauth-scopes"></p>
                        <button class="btn btn-gray" onclick="disconnectOAuth()">Disconnect</button>
                    </div>
                    <div id="oauth-not-connected">
                        <button class="btn btn-primary" id="btn-oauth" onclick="startOAuth()">Connect Google Calendar</button>
                    </div>
                    <div class="status info" id="oauth-status" style="display:none;"></div>
                </div>
                <div class="card">
                    <h2>3. Follow-up Flow</h2>
                    <div class="flow-step" id="step1"><div class="num">1</div><div>Submit lead details (form above)</div></div>
                    <div class="flow-step" id="step2"><div class="num">2</div><div>CALL-E SDK places outbound call to the phone</div></div>
                    <div class="flow-step" id="step3"><div class="num">3</div><div>CALL-E agent calls the lead, offers available 30-min slots (10 AM-6 PM) and confirms the timezone</div></div>
                    <div class="flow-step" id="step4"><div class="num">4</div><div>Google Calendar event created for the follow-up appointment</div></div>
                    <div class="flow-step" id="step5"><div class="num">5</div><div>Confirmation email sent to the lead</div></div>
                </div>
            </div>
        </div>

        <!-- Calendar event display -->
        <div class="card" id="calendar-card" style="display:none;">
            <h2>Last Created Calendar Event</h2>
            <div class="calendar-event">
                <h4 id="cal-title">Follow-up appointment</h4>
                <p id="cal-details"></p>
                <p><a id="cal-link" href="#" target="_blank">Open in Google Calendar</a></p>
            </div>
        </div>

    </div>

    <script>
        let callSid = null;
        let pollTimer = null;

        function apiToken() {
            return (document.getElementById('api-token') || {}).value || '';
        }

        function authHeaders(extra) {
            const h = Object.assign({ 'Authorization': 'Bearer ' + apiToken() }, extra || {});
            if (document.getElementById('live-confirm') && document.getElementById('live-confirm').checked) {
                h['X-Confirm-Live-Call'] = 'I understand this places a real phone call';
            }
            return h;
        }

        function submitLead() {
            const data = {
                name: document.getElementById('name').value,
                phone: buildE164(),
                email: document.getElementById('email').value,
                company: document.getElementById('company').value,
                your_company: document.getElementById('your-company').value,
                company_tz: document.getElementById('company-tz').value,
                consent: document.getElementById('lead-consent').checked ? 'yes' : 'no'
            };
            const statusBox = document.getElementById('call-status');
            const resultBox = document.getElementById('call-result');
            const btn = document.getElementById('btn-call');
            btn.disabled = true;
            statusBox.className = 'status info';
            statusBox.style.display = 'block';
            statusBox.textContent = 'Placing CALL-E outbound call...';
            setStep(1, 'active');

            fetch('/lead-submission', {
                method: 'POST',
                headers: authHeaders({ 'Content-Type': 'application/json' }),
                body: JSON.stringify(data)
            })
            .then(r => r.json().then(res => ({ res, ok: r.ok })))
            .then(({ res, ok }) => {
                resultBox.style.display = 'block';
                resultBox.textContent = JSON.stringify(res, null, 2);
                if (!ok || res.error) {
                    statusBox.className = 'status error';
                    statusBox.textContent = res.error || 'Request failed';
                    setStep(1, '');
                    return;
                }
                callSid = res.call_sid;
                statusBox.className = 'status success';
                statusBox.textContent = 'Call placed! SID: ' + res.call_sid + ' | Status: ' + res.call_status;
                setStep(2, 'done');
                setStep(3, 'active');
                statusBox.textContent = 'Call placed! Waiting for the lead to answer (SID: ' + res.call_sid + ')...';

                // Poll the real call status until terminal, then schedule at chosen time
                pollCallStatus();
            })
            .catch(err => {
                statusBox.className = 'status error';
                statusBox.textContent = 'Request failed: ' + err;
            })
            .finally(() => { btn.disabled = false; });
        }

        function pollCallStatus() {
            if (pollTimer) clearTimeout(pollTimer);
            fetch('/call-status/' + callSid, { headers: authHeaders() })
            .then(r => r.json())
            .then(res => {
                if (res.error) {
                    const statusBox = document.getElementById('call-status');
                    statusBox.className = 'status error';
                    statusBox.textContent = 'Error: ' + res.error;
                    return;
                }
                const statusBox = document.getElementById('call-status');
                const resultBox = document.getElementById('call-result');

                if (res.status === 'completed' || res.status === 'failed' || res.status === 'canceled') {
                    setStep(3, 'done');
                    resultBox.textContent = JSON.stringify(res, null, 2);

                    const wants = res.structured_result && res.structured_result.wants_appointment;
                    if (res.status === 'completed' && wants === 'yes') {
                        statusBox.className = 'status success';
                        statusBox.textContent = 'Lead accepted! Preferred: ' +
                            ((res.structured_result && res.structured_result.preferred_day) || 'unknown') + ' at ' +
                            ((res.structured_result && res.structured_result.preferred_time) || 'unknown') +
                            '. Scheduling appointment...';
                        setStep(4, 'active');
                        if (res.calendar && res.calendar.status === 'success') {
                            showCalendarEvent(res.calendar);
                            setStep(4, 'done');
                            setStep(5, 'active');
                            statusBox.textContent = 'Appointment scheduled: ' +
                                res.calendar.preferred_time + ' ' + res.calendar.start + ' | Event: ' + res.calendar.event_id;
                            if (res.email && res.email.status === 'success') {
                                statusBox.textContent += ' | Confirmation email sent to ' + res.email.to;
                            } else if (res.email && res.email.status === 'demo') {
                                statusBox.textContent += ' | Email logged to console (SMTP not configured)';
                            } else if (res.email && res.email.status === 'error') {
                                statusBox.className = 'status info';
                                statusBox.textContent += ' | Email failed: ' + res.email.message;
                            }
                            setTimeout(() => setStep(5, 'done'), 1000);
                        } else if (res.calendar && res.calendar.status === 'oauth_required') {
                            statusBox.className = 'status info';
                            statusBox.textContent = 'Call completed. Connect Google Calendar to complete scheduling.';
                        } else if (res.calendar && res.calendar.status === 'no_time') {
                            statusBox.className = 'status info';
                            statusBox.textContent = res.calendar.message;
                        } else if (res.calendar && res.calendar.status === 'error') {
                            statusBox.className = 'status error';
                            statusBox.textContent = 'Calendar error: ' + res.calendar.message;
                        }
                    } else {
                        statusBox.className = 'status info';
                        statusBox.textContent = 'Call ' + res.status + '. Appointment: ' + (wants || 'unknown');
                    }
                } else {
                    // Keep polling
                    statusBox.textContent = 'Call status: ' + res.status + ' - still ringing...';
                    pollTimer = setTimeout(pollCallStatus, 5000);
                }
            })
            .catch(err => {
                pollTimer = setTimeout(pollCallStatus, 5000);
            });
        }

        function startOAuth() {
            fetch('/oauth', { headers: authHeaders() })
            .then(r => r.json())
            .then(res => {
                if (res.url) {
                    window.location.href = res.url;
                } else if (res.error) {
                    const statusBox = document.getElementById('oauth-status');
                    statusBox.style.display = 'block';
                    statusBox.className = 'status error';
                    statusBox.textContent = 'Error: ' + res.error;
                }
            })
            .catch(err => {
                const statusBox = document.getElementById('oauth-status');
                statusBox.style.display = 'block';
                statusBox.className = 'status error';
                statusBox.textContent = 'Request failed: ' + err;
            });
        }

        function disconnectOAuth() {
            const statusBox = document.getElementById('oauth-status');
            statusBox.style.display = 'block';
            statusBox.className = 'status info';
            statusBox.textContent = 'Disconnecting...';
            fetch('/oauth-disconnect', { method: 'POST', headers: authHeaders() })
            .then(r => r.json())
            .then(res => {
                if (res.error) {
                    statusBox.className = 'status error';
                    statusBox.textContent = 'Error: ' + res.error;
                    return;
                }
                statusBox.className = 'status success';
                statusBox.textContent = 'Disconnected. You can reconnect anytime.';
                checkOAuthStatus();
            })
            .catch(err => {
                statusBox.className = 'status error';
                statusBox.textContent = 'Request failed: ' + err;
            });
        }

        function populateCountries() {
            const countries = [
                ['+91', 'India'], ['+1', 'US/Canada'], ['+44', 'UK'], ['+61', 'Australia'],
                ['+971', 'UAE'], ['+65', 'Singapore'], ['+86', 'China'], ['+81', 'Japan'],
                ['+92', 'Pakistan'], ['+66', 'Thailand'], ['+62', 'Indonesia'], ['+852', 'Hong Kong'],
                ['+49', 'Germany'], ['+33', 'France'], ['+34', 'Spain'], ['+39', 'Italy'],
                ['+7', 'Russia'], ['+20', 'Egypt'], ['+234', 'Nigeria'], ['+27', 'South Africa'],
                ['+55', 'Brazil'], ['+52', 'Mexico'], ['+57', 'Colombia'], ['+64', 'New Zealand'],
                ['+880', 'Bangladesh'], ['+94', 'Sri Lanka'], ['+977', 'Nepal'], ['+60', 'Malaysia'],
                ['+84', 'Vietnam'], ['+63', 'Philippines'], ['+90', 'Turkey'], ['+966', 'Saudi Arabia'],
                ['+973', 'Bahrain'], ['+974', 'Qatar'], ['+968', 'Oman'], ['+965', 'Kuwait']
            ];
            const select = document.getElementById('phone-country');
            for (const [code, name] of countries) {
                const opt = document.createElement('option');
                opt.value = code;
                opt.textContent = code + ' (' + name + ')';
                select.appendChild(opt);
            }
            select.value = '+91';
        }

        function buildE164() {
            const code = document.getElementById('phone-country').value;
            const number = document.getElementById('phone').value.replace(/[^0-9]/g, '');
            return number ? code + number : '';
        }

        let tzTimer = null;
        function updateLeadTz() {
            const display = document.getElementById('lead-tz-display');
            const phone = buildE164();
            if (phone.length < 10) {
                display.textContent = "Enter a phone number to auto-detect the lead's timezone.";
                return;
            }
            if (tzTimer) clearTimeout(tzTimer);
            tzTimer = setTimeout(() => {
                display.textContent = 'Detecting timezone...';
                fetch('/detect-timezone?phone=' + encodeURIComponent(phone), { headers: authHeaders() })
                .then(r => r.json())
                .then(res => {
                    if (res.timezone) {
                        display.textContent = 'Detected timezone: ' + (res.label || res.timezone);
                    } else if (res.error) {
                        display.textContent = res.error;
                    } else {
                        display.textContent = 'Timezone unknown - the meeting will use your timezone.';
                    }
                })
                .catch(() => {
                    display.textContent = 'Could not detect timezone.';
                });
            }, 400);
        }

        function setMode(mode) {
            const single = document.getElementById('single-mode');
            const bulk = document.getElementById('bulk-mode');
            const btnSingle = document.getElementById('mode-single');
            const btnBulk = document.getElementById('mode-bulk');
            if (mode === 'bulk') {
                single.style.display = 'none';
                bulk.style.display = 'block';
                btnSingle.className = 'btn mode-btn';
                btnSingle.style.background = '#a0aec0';
                btnBulk.className = 'btn mode-btn active';
                btnBulk.style.background = '#4a90d9';
            } else {
                bulk.style.display = 'none';
                single.style.display = 'block';
                btnBulk.className = 'btn mode-btn';
                btnBulk.style.background = '#a0aec0';
                btnSingle.className = 'btn mode-btn active';
                btnSingle.style.background = '#4a90d9';
            }
        }

        function checkOAuthStatus() {
            fetch('/oauth-status', { headers: authHeaders() })
            .then(r => r.json())
            .then(res => {
                const connected = document.getElementById('oauth-connected');
                const notConnected = document.getElementById('oauth-not-connected');
                const gateMsg = document.getElementById('gate-msg');
                const btnCall = document.getElementById('btn-call');
                const scopes = res.scopes || [];
                const hasCalendar = scopes.some(s => s.indexOf('calendar') !== -1);
                const hasGmail = scopes.some(s => s.indexOf('gmail') !== -1);

                if (res.connected && hasCalendar && hasGmail) {
                    document.getElementById('oauth-account').textContent = res.account || '(unknown account)';
                    document.getElementById('oauth-scopes').textContent =
                        'Granted scopes: ' + scopes.map(s => s.split('/').pop()).join(', ');
                    connected.style.display = 'block';
                    notConnected.style.display = 'none';
                    gateMsg.className = 'status success';
                    gateMsg.textContent = 'Google connected: ' + (res.account || 'your account') + '. You can now place calls - events go to your calendar.';
                    btnCall.disabled = false;
                } else if (res.connected) {
                    connected.style.display = 'block';
                    notConnected.style.display = 'none';
                    gateMsg.className = 'status info';
                    gateMsg.textContent = 'Google connected but missing permissions (calendar/gmail). Disconnect and reconnect to fix.';
                    btnCall.disabled = true;
                } else {
                    connected.style.display = 'none';
                    notConnected.style.display = 'block';
                    gateMsg.className = 'status info';
                    gateMsg.textContent = 'Please connect Google Calendar & Gmail first (right side) before placing calls.';
                    btnCall.disabled = true;
                }
            })
            .catch(() => {});
        }

        (function init() {
            setMode('single');
            populateCountries();
            document.getElementById('company-tz').value = Intl.DateTimeFormat().resolvedOptions().timeZone || 'UTC';
            document.getElementById('phone-country').addEventListener('change', updateLeadTz);
            document.getElementById('phone').addEventListener('input', updateLeadTz);
            updateLeadTz();
            checkOAuthStatus();
            const params = new URLSearchParams(window.location.search);
            if (params.get('oauth') === 'connected') {
                const statusBox = document.getElementById('oauth-status');
                statusBox.style.display = 'block';
                statusBox.className = 'status success';
                statusBox.textContent = 'Google account connected successfully!';
                history.replaceState({}, '', window.location.pathname);
            } else if (params.get('oauth') === 'error') {
                const statusBox = document.getElementById('oauth-status');
                statusBox.style.display = 'block';
                statusBox.className = 'status error';
                statusBox.textContent = 'OAuth failed: ' + (params.get('message') || 'unknown error');
                history.replaceState({}, '', window.location.pathname);
            }
        })();

        function setStep(n, state) {
            const el = document.getElementById('step' + n);
            el.className = 'flow-step ' + state;
        }

        function showCalendarEvent(cal) {
            document.getElementById('calendar-card').style.display = 'block';
            const start = new Date(cal.start);
            document.getElementById('cal-details').textContent =
                'Start: ' + start.toLocaleString() +
                ' | End: ' + new Date(cal.end).toLocaleString() +
                ' | Event ID: ' + cal.event_id;
            if (cal.htmlLink) {
                document.getElementById('cal-link').href = cal.htmlLink;
            }
        }

        let currentBatchId = null;
        let batchPollTimer = null;

        function uploadBatch() {
            const fileInput = document.getElementById('batch-file');
            if (!fileInput.files || !fileInput.files[0]) {
                document.getElementById('batch-status').style.display = 'block';
                document.getElementById('batch-status').className = 'status error';
                document.getElementById('batch-status').textContent = 'Please choose a file first.';
                return;
            }
            const fd = new FormData();
            fd.append('file', fileInput.files[0]);
            const statusBox = document.getElementById('batch-status');
            const btn = document.getElementById('btn-batch');
            btn.disabled = true;
            statusBox.style.display = 'block';
            statusBox.className = 'status info';
            statusBox.textContent = 'Uploading and starting batch...';

            fetch('/batch-upload', { method: 'POST', headers: authHeaders(), body: fd })
            .then(r => r.json().then(res => ({ res, ok: r.ok })))
            .then(({ res, ok }) => {
                if (!ok || res.error) {
                    statusBox.className = 'status error';
                    statusBox.textContent = res.error || 'Upload failed';
                    btn.disabled = false;
                    return;
                }
                currentBatchId = res.batch_id;
                statusBox.className = 'status success';
                statusBox.textContent = res.message;
                document.getElementById('btn-batch-stop').style.display = 'inline-block';
                pollBatch();
            })
            .catch(err => {
                statusBox.className = 'status error';
                statusBox.textContent = 'Request failed: ' + err;
                btn.disabled = false;
            });
        }

        function pollBatch() {
            if (!currentBatchId) return;
            if (batchPollTimer) clearTimeout(batchPollTimer);
            fetch('/batch-status/' + currentBatchId, { headers: authHeaders() })
            .then(r => r.json())
            .then(res => {
                if (res.error) return;
                const rows = res.rows || [];
                let html = '<b>Batch progress: ' + res.done_count + ' / ' + res.total + '</b>\\n\\n';
                html += 'Name'.padEnd(18) + 'Status'.padEnd(12) + 'Appointment'.padEnd(22) + 'Details\\n';
                for (const row of rows) {
                    const details = row.error || row.event_id || row.call_sid || '';
                    html += (row.name || '?').padEnd(18) + (row.status || 'pending').padEnd(12) +
                            (row.appointment || '-').padEnd(22) + details + '\\n';
                }
                document.getElementById('batch-result').textContent = html;
                document.getElementById('batch-result').style.display = 'block';

                if (!res.running) {
                    document.getElementById('btn-batch-stop').style.display = 'none';
                    document.getElementById('btn-batch-download').style.display = 'inline-block';
                    document.getElementById('btn-batch').disabled = false;
                    const statusBox = document.getElementById('batch-status');
                    statusBox.className = 'status success';
                    statusBox.textContent = 'Batch complete! Download the updated Excel file.';
                    return;
                }
                batchPollTimer = setTimeout(pollBatch, 5000);
            })
            .catch(() => {
                batchPollTimer = setTimeout(pollBatch, 5000);
            });
        }

        function stopBatch() {
            if (!currentBatchId) return;
            fetch('/batch-stop/' + currentBatchId, { method: 'POST', headers: authHeaders() })
            .then(r => r.json())
            .then(res => {
                document.getElementById('batch-status').textContent = 'Stopping after current call...';
            })
            .catch(() => {});
        }

        function downloadBatch() {
            if (!currentBatchId) return;
            fetch('/batch-download/' + currentBatchId, { headers: authHeaders() })
            .then(r => {
                if (!r.ok) return r.json().then(res => { throw new Error(res.error || 'Download failed'); });
                return r.blob();
            })
            .then(blob => {
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = 'batch_' + currentBatchId + '.xlsx';
                document.body.appendChild(a);
                a.click();
                a.remove();
                URL.revokeObjectURL(url);
            })
            .catch(err => {
                document.getElementById('batch-status').textContent = 'Download error: ' + err.message;
            });
        }
    </script>
</body>
</html>
"""


# ============================================================
# GOOGLE CALENDAR OAUTH & INTEGRATION
# ============================================================

def get_google_credentials(session_id=None):
    """Get valid Google OAuth credentials for the calling session.

    Tokens are stored per-session in server memory by default. Disk
    persistence / env restore happens only when PERSIST_GOOGLE_TOKEN=true
    (explicit operator opt-in). Returns None when unavailable.
    """
    sid = session_id or current_session_id()
    creds = None

    stored_json = oauth_tokens.get(sid)
    if stored_json:
        try:
            creds = Credentials.from_authorized_user_info(stored_json)
        except Exception:
            creds = None
    if creds is None and PERSIST_TOKEN:
        if os.path.exists(CREDENTIALS_FILE):
            try:
                creds = Credentials.from_authorized_user_file(CREDENTIALS_FILE)
            except Exception:
                creds = None
        elif os.environ.get("GOOGLE_TOKEN_JSON") and not os.path.exists(DISCONNECT_FLAG):
            # Operator-configured restore (ephemeral-disk hosts, e.g. Render).
            # A Disconnect click drops the flag file so the user stays disconnected.
            try:
                creds = Credentials.from_authorized_user_info(
                    json.loads(os.environ["GOOGLE_TOKEN_JSON"])
                )
                save_credentials(creds, session_id=sid)
            except Exception as e:
                print(f"GOOGLE_TOKEN_JSON restore failed: {e}")
                creds = None

    # Refresh when expired, fail closed otherwise.
    if creds and not creds.valid:
        if creds.expired and creds.refresh_token:
            try:
                from google.auth.transport.requests import Request
                creds.refresh(Request())
            except Exception as e:
                print(f"Google credential refresh failed: {e}")
                return None
        else:
            return None

    if creds:
        oauth_tokens[sid] = json.loads(creds.to_json())
    return creds


def save_credentials(creds, session_id=None):
    """Save credentials for a session; persist to disk only on explicit opt-in."""
    sid = session_id or current_session_id()
    oauth_tokens[sid] = json.loads(creds.to_json())
    if PERSIST_TOKEN:
        with open(CREDENTIALS_FILE, "w") as f:
            f.write(creds.to_json())


def parse_preferred_time(preferred_time):
    """Parse a HH:MM time string to a datetime.time, or None."""
    try:
        return datetime.strptime(preferred_time, "%H:%M").time()
    except (ValueError, TypeError):
        return None


# Fallback: primary timezone per region (used when the number is not
# recognized by phonenumbers' timezone table, e.g. placeholder/invalid numbers)
TZ_FRIENDLY = {
    # India / South Asia
    "Asia/Kolkata": "IST (India Standard Time)",
    "Asia/Colombo": "SLST (Sri Lanka Standard Time)",
    "Asia/Karachi": "PKT (Pakistan Standard Time)",
    "Asia/Dhaka": "BST (Bangladesh Standard Time)",
    "Asia/Kathmandu": "NPT (Nepal Time)",
    "Asia/Kabul": "AFT (Afghanistan Time)",
    # Gulf / Middle East
    "Asia/Dubai": "GST (Gulf Standard Time)",
    "Asia/Muscat": "GST (Gulf Standard Time)",
    "Asia/Riyadh": "AST (Arabia Standard Time)",
    "Asia/Qatar": "AST (Arabia Standard Time)",
    "Asia/Kuwait": "AST (Arabia Standard Time)",
    "Asia/Bahrain": "AST (Arabia Standard Time)",
    "Asia/Tehran": "IRST (Iran Standard Time)",
    "Asia/Baghdad": "AST (Arabia Standard Time)",
    "Asia/Amman": "EET (Eastern European Time)",
    "Asia/Beirut": "EET (Eastern European Time)",
    "Asia/Jerusalem": "IST (Israel Standard Time)",
    # Southeast / East Asia
    "Asia/Singapore": "SGT (Singapore Time)",
    "Asia/Shanghai": "CST (China Standard Time)",
    "Asia/Hong_Kong": "HKT (Hong Kong Time)",
    "Asia/Taipei": "CST (China Standard Time)",
    "Asia/Tokyo": "JST (Japan Standard Time)",
    "Asia/Seoul": "KST (Korea Standard Time)",
    "Asia/Bangkok": "ICT (Indochina Time)",
    "Asia/Ho_Chi_Minh": "ICT (Indochina Time)",
    "Asia/Phnom_Penh": "ICT (Indochina Time)",
    "Asia/Vientiane": "ICT (Indochina Time)",
    "Asia/Yangon": "MMT (Myanmar Time)",
    "Asia/Jakarta": "WIB (Western Indonesia Time)",
    "Asia/Kuala_Lumpur": "MYT (Malaysia Time)",
    "Asia/Manila": "PHT (Philippine Time)",
    "Asia/Ulaanbaatar": "ULAT (Ulaanbaatar Time)",
    "Asia/Tashkent": "UZT (Uzbekistan Time)",
    "Asia/Almaty": "ALMT (Almaty Time)",
    "Asia/Baku": "AZT (Azerbaijan Time)",
    "Asia/Tbilisi": "GET (Georgia Standard Time)",
    "Asia/Yerevan": "AMT (Armenia Time)",
    # Europe
    "Europe/London": "GMT (British Time)",
    "Europe/Dublin": "GMT (Irish Time)",
    "Europe/Paris": "CET (Central European Time)",
    "Europe/Berlin": "CET (Central European Time)",
    "Europe/Madrid": "CET (Central European Time)",
    "Europe/Rome": "CET (Central European Time)",
    "Europe/Amsterdam": "CET (Central European Time)",
    "Europe/Brussels": "CET (Central European Time)",
    "Europe/Zurich": "CET (Central European Time)",
    "Europe/Vienna": "CET (Central European Time)",
    "Europe/Stockholm": "CET (Central European Time)",
    "Europe/Oslo": "CET (Central European Time)",
    "Europe/Copenhagen": "CET (Central European Time)",
    "Europe/Warsaw": "CET (Central European Time)",
    "Europe/Prague": "CET (Central European Time)",
    "Europe/Lisbon": "WET (Western European Time)",
    "Europe/Helsinki": "EET (Eastern European Time)",
    "Europe/Athens": "EET (Eastern European Time)",
    "Europe/Kiev": "EET (Eastern European Time)",
    "Europe/Moscow": "MSK (Moscow Standard Time)",
    "Europe/Istanbul": "TRT (Turkey Time)",
    # Americas
    "America/New_York": "ET (Eastern Time)",
    "America/Toronto": "ET (Eastern Time)",
    "America/Jamaica": "ET (Eastern Time)",
    "America/Chicago": "CT (Central Time)",
    "America/Mexico_City": "CT (Central Time)",
    "America/Denver": "MT (Mountain Time)",
    "America/Los_Angeles": "PT (Pacific Time)",
    "America/Sao_Paulo": "BRT (Brasilia Time)",
    "America/Bogota": "COT (Colombia Time)",
    "America/Lima": "PET (Peru Time)",
    "America/Caracas": "VET (Venezuela Time)",
    "America/Santiago": "CLT (Chile Time)",
    "America/Argentina/Buenos_Aires": "ART (Argentina Time)",
    "America/Barbados": "AST (Atlantic Standard Time)",
    "America/Port_of_Spain": "AST (Atlantic Standard Time)",
    "America/St_Vincent": "AST (Atlantic Standard Time)",
    # Africa
    "Africa/Cairo": "EET (Eastern European Time)",
    "Africa/Lagos": "WAT (West Africa Time)",
    "Africa/Johannesburg": "SAST (South Africa Standard Time)",
    "Africa/Nairobi": "EAT (East Africa Time)",
    "Africa/Addis_Ababa": "EAT (East Africa Time)",
    "Africa/Accra": "GMT (Greenwich Mean Time)",
    "Africa/Casablanca": "WET (Western European Time)",
    # Oceania
    "Australia/Sydney": "AET (Australian Eastern Time)",
    "Australia/Melbourne": "AET (Australian Eastern Time)",
    "Australia/Brisbane": "AET (Australian Eastern Time)",
    "Australia/Adelaide": "ACT (Australian Central Time)",
    "Australia/Perth": "AWST (Australian Western Standard Time)",
    "Pacific/Auckland": "NZT (New Zealand Time)",
}


def tz_friendly_name(tz):
    """Map an IANA timezone to a human-friendly label, e.g. Asia/Kolkata -> IST (India Standard Time)."""
    if not tz:
        return None
    return TZ_FRIENDLY.get(tz, tz)


REGION_TZ = {
    "US": "America/New_York", "CA": "America/Toronto", "MX": "America/Mexico_City",
    "BR": "America/Sao_Paulo", "AR": "America/Argentina/Buenos_Aires", "CL": "America/Santiago",
    "CO": "America/Bogota", "PE": "America/Lima", "VE": "America/Caracas",
    "GB": "Europe/London", "IE": "Europe/Dublin", "FR": "Europe/Paris", "DE": "Europe/Berlin",
    "ES": "Europe/Madrid", "IT": "Europe/Rome", "PT": "Europe/Lisbon", "NL": "Europe/Amsterdam",
    "BE": "Europe/Brussels", "CH": "Europe/Zurich", "AT": "Europe/Vienna", "SE": "Europe/Stockholm",
    "NO": "Europe/Oslo", "DK": "Europe/Copenhagen", "FI": "Europe/Helsinki", "PL": "Europe/Warsaw",
    "GR": "Europe/Athens", "CZ": "Europe/Prague", "UA": "Europe/Kiev", "RU": "Europe/Moscow",
    "TR": "Europe/Istanbul", "IL": "Asia/Jerusalem", "SA": "Asia/Riyadh", "AE": "Asia/Dubai",
    "QA": "Asia/Qatar", "KW": "Asia/Kuwait", "BH": "Asia/Bahrain", "OM": "Asia/Muscat",
    "JO": "Asia/Amman", "LB": "Asia/Beirut", "IQ": "Asia/Baghdad", "IR": "Asia/Tehran",
    "EG": "Africa/Cairo", "NG": "Africa/Lagos", "ZA": "Africa/Johannesburg", "MA": "Africa/Casablanca",
    "KE": "Africa/Nairobi", "ET": "Africa/Addis_Ababa", "GH": "Africa/Accra",
    "IN": "Asia/Kolkata", "PK": "Asia/Karachi", "BD": "Asia/Dhaka", "LK": "Asia/Colombo",
    "NP": "Asia/Kathmandu", "AF": "Asia/Kabul", "MM": "Asia/Yangon", "TH": "Asia/Bangkok",
    "ID": "Asia/Jakarta", "VN": "Asia/Ho_Chi_Minh", "MY": "Asia/Kuala_Lumpur", "SG": "Asia/Singapore",
    "PH": "Asia/Manila", "KH": "Asia/Phnom_Penh", "LA": "Asia/Vientiane",
    "CN": "Asia/Shanghai", "TW": "Asia/Taipei", "HK": "Asia/Hong_Kong", "JP": "Asia/Tokyo",
    "KR": "Asia/Seoul", "MN": "Asia/Ulaanbaatar", "UZ": "Asia/Tashkent", "KZ": "Asia/Almaty",
    "GE": "Asia/Tbilisi", "AZ": "Asia/Baku", "AM": "Asia/Yerevan",
    "AU": "Australia/Sydney", "NZ": "Pacific/Auckland",
    "VC": "America/St_Vincent", "BB": "America/Barbados", "JM": "America/Jamaica", "TT": "America/Port_of_Spain",
}
COUNTRY_CODE_TZ = {
    1: "America/New_York", 44: "Europe/London", 61: "Australia/Sydney", 64: "Pacific/Auckland",
    971: "Asia/Dubai", 966: "Asia/Riyadh", 974: "Asia/Qatar", 965: "Asia/Kuwait",
    968: "Asia/Muscat", 973: "Asia/Bahrain", 49: "Europe/Berlin", 33: "Europe/Paris",
    34: "Europe/Madrid", 39: "Europe/Rome", 7: "Europe/Moscow", 20: "Africa/Cairo",
    234: "Africa/Lagos", 27: "Africa/Johannesburg", 55: "America/Sao_Paulo", 52: "America/Mexico_City",
    57: "America/Bogota", 92: "Asia/Karachi", 880: "Asia/Dhaka", 94: "Asia/Colombo",
    977: "Asia/Kathmandu", 66: "Asia/Bangkok", 62: "Asia/Jakarta", 84: "Asia/Ho_Chi_Minh",
    63: "Asia/Manila", 60: "Asia/Kuala_Lumpur", 86: "Asia/Shanghai", 81: "Asia/Tokyo",
    82: "Asia/Seoul", 852: "Asia/Hong_Kong", 886: "Asia/Taipei", 90: "Europe/Istanbul",
    91: "Asia/Kolkata", 65: "Asia/Singapore",
}


def get_lead_timezone(phone):
    """Derive the lead's IANA timezone from their phone number, or None."""
    try:
        import phonenumbers
        from phonenumbers import timezone as pn_timezone

        num = phonenumbers.parse(phone, None)
        zones = pn_timezone.time_zones_for_number(num)
        for z in zones:
            if z and z != "Etc/Unknown":
                # Normalize old-style zone names (e.g. Asia/Calcutta -> Asia/Kolkata)
                return z.replace("Calcutta", "Kolkata")

        # Fallback 1: region's primary timezone
        region = phonenumbers.region_code_for_number(num)
        if region and region in REGION_TZ:
            return REGION_TZ[region]

        # Fallback 2: country calling code's primary timezone
        return COUNTRY_CODE_TZ.get(num.country_code)
    except Exception:
        return None


def convert_lead_time_to_company(lead_tz, company_tz, day_str, time_str, base_date):
    """Convert a lead's local (day, HH:MM) into the company timezone.

    Returns (start_dt, lead_tz_label, company_tz_label) as aware datetimes/strings,
    or (None, ...) when conversion fails (fall back to naive company tz).
    """
    try:
        from zoneinfo import ZoneInfo

        t = parse_preferred_time(time_str)
        if t is None:
            return None, None, None

        # If we don't know the lead's timezone, interpret the time in the
        # company timezone (no conversion needed).
        if not lead_tz:
            naive = datetime.combine(base_date, t)
            aware_company = naive.replace(tzinfo=ZoneInfo(company_tz or "UTC"))
            return aware_company, None, company_tz

        naive = datetime.combine(base_date, t)
        if day_str == "today":
            # If the naive time already passed today, keep as-is for demo simplicity
            pass
        aware_lead = naive.replace(tzinfo=ZoneInfo(lead_tz))
        aware_utc = aware_lead.astimezone(ZoneInfo("UTC"))
        aware_company = aware_utc.astimezone(ZoneInfo(company_tz))
        return aware_company, lead_tz, company_tz
    except Exception:
        return None, None, None


# ============================================================
# AVAILABILITY / SLOT LOGIC (30-min slots, 10 AM - 6 PM)
# ============================================================

BUSINESS_START_HOUR = 10
BUSINESS_END_HOUR = 18
SLOT_MINUTES = 30


def get_calendar_busy(service, day, tz):
    """Return busy intervals [(start, end), ...] for `day` on the primary calendar,
    as aware datetimes. Works with the calendar.events scope (events().list)."""
    tzinfo = ZoneInfo(tz or "UTC")
    start = datetime.combine(day, datetime.min.time()).replace(tzinfo=tzinfo)
    end = start + timedelta(days=1)
    resp = service.events().list(
        calendarId="primary",
        timeMin=start.isoformat(),
        timeMax=end.isoformat(),
        singleEvents=True,
    ).execute()
    busy = []
    for item in resp.get("items", []):
        s = item.get("start", {})
        e = item.get("end", {})
        if s.get("dateTime") and e.get("dateTime"):
            busy.append((datetime.fromisoformat(s["dateTime"]), datetime.fromisoformat(e["dateTime"])))
        elif s.get("date") and e.get("date"):
            # All-day events block the whole day
            d1 = datetime.fromisoformat(s["date"]).replace(tzinfo=tzinfo)
            d2 = datetime.fromisoformat(e["date"]).replace(tzinfo=tzinfo)
            busy.append((d1, d2))
    return busy


def get_available_slots(service, day, tz, lead_tz=None):
    """Compute free 30-minute business-hour slots (10:00-18:00) for `day`.

    Returns slot dicts: {"start": aware datetime (company tz),
    "company": "HH:MM", "lead": "HH:MM"}. `lead` shows the same instant
    in the lead's timezone when lead_tz is given. Past slots are skipped.
    """
    tzinfo = ZoneInfo(tz or "UTC")
    day_start = datetime.combine(day, datetime.min.time()).replace(tzinfo=tzinfo)
    now = datetime.now(tzinfo)

    busy = get_calendar_busy(service, day, tz or "UTC")

    slots = []
    t = day_start.replace(hour=BUSINESS_START_HOUR, minute=0)
    end_of_day = day_start.replace(hour=BUSINESS_END_HOUR, minute=0)
    while t + timedelta(minutes=SLOT_MINUTES) <= end_of_day:
        slot_end = t + timedelta(minutes=SLOT_MINUTES)
        conflict = any(s < slot_end and e > t for s, e in busy)
        if not conflict and t >= now:
            lead_time = t.astimezone(ZoneInfo(lead_tz)) if lead_tz else t
            slots.append({
                "start": t,
                "company": t.strftime("%H:%M"),
                "lead": lead_time.strftime("%H:%M"),
            })
        t += timedelta(minutes=SLOT_MINUTES)
    return slots


def format_slots_for_message(slots):
    if not slots:
        return "no available slots at the moment"
    return ", ".join(s["lead"] for s in slots)


def schedule_google_calendar(name, email, phone, company, your_company=None, preferred_day="tomorrow", preferred_time=None, call_sid=None, lead_tz=None, company_tz=None):
    """Schedule Google Calendar appointment after CALL-E call.

    The event is created in the caller's (connected company account) calendar.
    `company` is the lead's company; `your_company` is the caller's company.

    Timezone handling: the lead's preferred (day, HH:MM) is interpreted in the
    lead's timezone (derived from phone number), converted to the company
    timezone, and the event is stored with the company timezone. If either
    timezone is unknown, the company timezone (or UTC) is used as-is.
    """
    try:
        creds = get_google_credentials()
        
        if not creds:
            # Return demo result - user needs to complete OAuth
            return {
                "status": "oauth_required",
                "message": "Google Calendar scheduling requires OAuth completion. Run oauth_setup.py to authorize.",
                "action_required": "complete_oauth"
            }
        
        # Build the Google Calendar service
        service = build("calendar", "v3", credentials=creds)
        
        # Determine the appointment start time
        now = datetime.utcnow()
        if preferred_day == "today":
            base_day = now.date()
        else:
            base_day = now.date() + timedelta(days=1)  # tomorrow

        company_tz = company_tz or "UTC"
        start_time = None
        tz_note = ""

        if preferred_time:
            if lead_tz and lead_tz != company_tz:
                # Convert lead's local time into the company timezone
                aware, _, _ = convert_lead_time_to_company(
                    lead_tz, company_tz, preferred_day, preferred_time, base_day
                )
                if aware is not None:
                    start_time = aware
                    tz_note = f"({preferred_time} {lead_tz} = {aware.strftime('%H:%M')} {company_tz})"
            if start_time is None:
                # Fallback: interpret time directly in company tz
                parsed = parse_preferred_time(preferred_time)
                if parsed:
                    start_time = datetime.combine(base_day, parsed).replace(tzinfo=ZoneInfo(company_tz))
                    tz_note = f"({preferred_time} {company_tz})"

        if start_time is None:
            return {
                "status": "error",
                "message": "No valid confirmed time available - nothing was booked.",
            }

        # Re-check the chosen slot is still free (another lead/meeting may have
        # taken it since the call). Fail closed: if the lookup fails, or the
        # confirmed slot is no longer free, refuse to book. Never silently move
        # to a different time the lead did not confirm.
        try:
            busy = get_calendar_busy(service, start_time.date(), company_tz)
        except Exception as e:
            print(f"Conflict check failed: {e}")
            return {
                "status": "error",
                "message": f"Availability lookup failed ({e}); the appointment was NOT booked. Retry or book manually.",
            }
        if any(s < start_time + timedelta(minutes=SLOT_MINUTES) and e > start_time for s, e in busy):
            return {
                "status": "error",
                "message": (
                    f"The confirmed slot ({start_time.strftime('%H:%M')} {company_tz}) is no longer free; "
                    "the appointment was NOT booked at a different time. Offer the free slots and re-confirm with the lead."
                ),
            }

        end_time = start_time + timedelta(minutes=SLOT_MINUTES)  # 30-min appointment
        
        caller = your_company or "Company"
        lead_org = company or "Lead's Company"
        summary = f"Follow-up with {name} ({caller})"
        event = {
            "summary": summary,
            "description": (
                f"Scheduled follow-up call with {name} from {lead_org}.\n"
                f"Caller: {caller}\n"
                f"Lead interest captured via CALL-E automation.\n"
                f"Phone: {mask_phone(phone)}\n"
                f"Lead email: {email}\n"
                f"Call SID: {call_sid or 'n/a'}\n"
                f"Time: {start_time.isoformat()} {tz_note}"
            ),
            "start": {
                "dateTime": start_time.isoformat(),
                "timeZone": company_tz,
            },
            "end": {
                "dateTime": end_time.isoformat(),
                "timeZone": company_tz,
            },
            "attendees": [
                {"email": email} if email else {"displayName": name}
            ],
        }
        
        event = service.events().insert(calendarId="primary", body=event).execute()
        return {
            "status": "success",
            "event_id": event.get("id"),
            "htmlLink": event.get("htmlLink"),
            "start": event.get("start", {}).get("dateTime"),
            "end": event.get("end", {}).get("dateTime"),
            "preferred_time": preferred_time,
            "timezone": company_tz,
            "tz_note": tz_note,
        }
        
    except Exception as e:
        print(f"Google Calendar Error: {e}")
        return {"status": "error", "message": str(e)}


# ============================================================
# CONFIRMATION EMAIL
# ============================================================

def send_confirmation_email(name, email, company, event, phone="", your_company=None):
    """Send appointment confirmation email to the lead via the Gmail API.

    Uses the same Google OAuth credentials as Calendar (adds gmail.send scope).
    Falls back to a logged/demo mode when OAuth is unavailable.
    Returns a dict describing the send result.
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from base64 import urlsafe_b64encode

    start = event.get("start", "")
    end = event.get("end", "")
    html_link = event.get("htmlLink", "")
    tz_note = event.get("tz_note", "")
    caller = your_company or "Company"
    lead_org = company or "your team"

    subject = f"Appointment Confirmed: Follow-up with {caller}"
    tz_line = f"\n  Time note: {tz_note}" if tz_note else ""
    body = f"""Hi {name},

Thank you for scheduling a follow-up appointment with {caller}.

Details:
  Date/Time: {start} - {end}
  Duration: 30 minutes{tz_line}

You can view or add it to your calendar here:
  {html_link}

If you need to reschedule, simply reply to this email.

- {caller} Team
"""
    try:
        creds = get_google_credentials()
        if not creds:
            # Demo mode: no valid OAuth for Gmail
            print(f"[EMAIL-DEMO] To: {email} | Subject: {subject}\n{body}")
            return {
                "status": "demo",
                "message": "Google OAuth not authorized for Gmail - email logged to server console.",
                "to": email,
            }

        has_gmail_scope = any(
            "gmail.send" in s for s in (getattr(creds, "scopes", None) or [])
        )
        if not has_gmail_scope:
            print(f"[EMAIL-DEMO] To: {email} | Subject: {subject}\n{body}")
            return {
                "status": "demo",
                "message": (
                    "Token lacks gmail.send scope - visit /oauth to re-authorize, "
                    "then the confirmation email will be sent via Gmail API."
                ),
                "to": email,
            }

        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = "me"
        msg["To"] = email
        msg.attach(MIMEText(body, "plain"))

        raw = urlsafe_b64encode(msg.as_bytes()).decode("ascii")
        service = build("gmail", "v1", credentials=creds)
        sent = service.users().messages().send(
            userId="me",
            body={"raw": raw},
        ).execute()

        return {
            "status": "success",
            "to": email,
            "subject": subject,
            "message_id": sent.get("id"),
        }

    except Exception as e:
        print(f"Email Error: {e}")
        return {"status": "error", "message": str(e), "to": email}


# ============================================================
# CALL-E CALL FUNCTION
# ============================================================

def generate_call_message(name, your_company, lead_company=None, lead_tz=None, slots=None, slot_day="tomorrow"):
    """Generate a natural-language CALL-E task for the follow-up call."""
    context = ""
    if lead_company:
        context = f" The lead works at {lead_company}."
    tz_label = tz_friendly_name(lead_tz) if lead_tz else None
    tz_line = ""
    if tz_label:
        tz_line = f"""
The lead's phone number indicates they are in the {tz_label} timezone. 
When they give a preferred time, repeat it back and explicitly confirm the time is in {tz_label}. 
Wait for their verbal confirmation before recording the time."""
    slots_line = ""
    if slots:
        slot_times = format_slots_for_message(slots)
        tz_clause = f" in {tz_label}" if tz_label else ""
        slots_line = f"""
The available appointment slots for {slot_day}{tz_clause} are: {slot_times}.
IMPORTANT: Only offer times from this list. If the lead requests a time that is not in the list, tell them that time is already booked, and offer the available slots instead. Confirm the chosen slot with them before recording it."""
    return f"""Call the lead {name} on behalf of {your_company}.{context}
AI DISCLOSURE (mandatory, first): Begin the call by clearly stating:
'Hello {name}, this is {your_company}'s AI voice assistant returning your follow-up request. I am an AI and this call is automated.'
If the lead is surprised, objects to speaking with an AI, or asks to end the call, apologize, end the call politely, and record wants_appointment=no - do not continue.
Then, if they are willing to continue, ask if they would like to schedule a quick 30-minute appointment on our calendar. 
If they accept, ask which day (today or tomorrow) and what time works best for them in their local timezone.{slots_line}
After they give a time, repeat it back and ask them to confirm the day, time, and that it is in their local timezone.
Record the result only after the lead confirms the time.{tz_line}"""


CALL_E_RESULT_SCHEMA = {
    "type": "object",
    "required": ["wants_appointment", "time_confirmed"],
    "properties": {
        "wants_appointment": {
            "type": "string",
            "enum": ["yes", "no", "unknown"],
            "description": "Use yes when the lead accepts a follow-up appointment, no when they decline, unknown when unclear.",
        },
        "preferred_day": {
            "type": "string",
            "enum": ["today", "tomorrow", "unknown"],
            "description": "The day the lead prefers for the appointment. Use unknown when not specified.",
        },
        "preferred_time": {
            "type": "string",
            "description": "The time the lead prefers for the appointment in HH:MM 24-hour format, e.g. 10:00. Use unknown when not specified.",
        },
        "time_confirmed": {
            "type": "string",
            "enum": ["yes", "no", "unknown"],
            "description": "Use yes only when the lead explicitly confirmed the day, time, and timezone. Use no if they disagreed or were unsure. Use unknown when not confirmed.",
        },
        "timezone": {
            "type": "string",
            "description": "The timezone the lead confirmed their time is in, as an IANA name such as Asia/Kolkata or America/New_York. Use unknown when not specified.",
        },
        "lead_reason": {
            "type": "string",
            "description": "Short reason the lead gave for the follow-up appointment, or unknown.",
        },
    },
    "additionalProperties": False,
}


def lead_idempotency_key(name, email, phone, company, your_company, day):
    """Deterministic provider idempotency key: an exact retry never re-calls."""
    import hashlib
    digest = hashlib.sha256(
        "|".join([name or "", email or "", phone or "", company or "", your_company or "", day or ""]).encode()
    ).hexdigest()
    return "lead-" + digest[:40]


def place_call_e_call(to_number, message, result_schema=None, idempotency_key=None):
    """Place an outbound CALL-E call, but only when live mode is explicitly enabled.

    Without CALLE_LIVE_CALLS_ENABLED=true the result is a clearly marked
    simulation (status 'simulated') and no call is ever placed.
    """
    if not LIVE_CALLS_ENABLED:
        return {
            "sid": None,
            "status": "simulated",
            "mode": "simulated",
            "simulated": True,
            "error": "Live calls are disabled server-side. Set CALLE_LIVE_CALLS_ENABLED=true to enable.",
        }
    try:
        result = call_e_client.call(
            to_number=to_number,
            message=message,
            result_schema=result_schema,
            idempotency_key=idempotency_key,
        )
        return result
    except Exception as e:
        print(f"CALL-E Error: {e}")
        return {"sid": None, "status": "failed", "error": str(e), "mode": "real"}


# ============================================================
# FLASK ROUTES
# ============================================================

@app.route("/")
def index():
    return render_template_string(DEMO_HTML)


@app.route("/lead-submission", methods=["POST"])
def lead_submission():
    """Handle lead form submission - triggers CALL-E call flow."""
    try:
        auth_err = require_token()
        if auth_err:
            return auth_err

        # Final opt-in for live calls: every live submission must carry the
        # confirmation header. Simulations need no header.
        if LIVE_CALLS_ENABLED:
            confirm = request.headers.get("X-Confirm-Live-Call", "")
            if confirm != "I understand this places a real phone call":
                return jsonify({
                    "error": "Live calls require the confirmation header "
                             "'X-Confirm-Live-Call: I understand this places a real phone call'.",
                }), 403

        lead_data = request.get_json()
        if not lead_data:
            return jsonify({"error": "No JSON data received"}), 400

        name = lead_data.get("name", "").strip() or "Valued Customer"
        email = lead_data.get("email", "").strip()
        phone = validate_e164(lead_data.get("phone", ""))
        if not phone:
            return jsonify({"error": "Invalid phone number (must be a valid E.164 number)."}), 400
        company = lead_data.get("company", "")  # lead's company
        your_company = lead_data.get("your_company", "Company")  # caller's company
        company_tz = lead_data.get("company_tz", "UTC") or "UTC"  # caller's timezone
        lead_tz = get_lead_timezone(phone)  # derived from phone number

        # Contact authorization: live calls are only placed for leads whose
        # consent is recorded in the request (and stamped server-side).
        consent = (lead_data.get("consent") or "").strip().lower()
        if LIVE_CALLS_ENABLED and consent != "yes":
            return jsonify({
                "error": "Live calls require recorded lead consent: set consent=yes (the lead must have authorized this follow-up call).",
            }), 400

        # Preview mode (live disabled) needs no Google connection: the result
        # is simulated and nothing is dialed or booked.
        creds = None
        if LIVE_CALLS_ENABLED:
            creds = get_google_credentials()
            has_gmail = any(
                "gmail.send" in s for s in (getattr(creds, "scopes", None) or [])
            )
            has_calendar = any(
                "calendar" in s for s in (getattr(creds, "scopes", None) or [])
            )
            if not creds or not has_gmail or not has_calendar:
                return jsonify({
                    "error": "Google Calendar & Gmail are not connected. Visit /oauth to authorize, then retry.",
                    "status": "oauth_required",
                }), 403

        # Pre-compute tomorrow's free 30-min slots (10 AM-6 PM) so the agent
        # can offer only available times and handle "that time is taken".
        # Live mode fails closed BEFORE dialing: if the calendar lookup fails
        # or no calendar-confirmed slot is free, no call is placed at all.
        slots = None
        if LIVE_CALLS_ENABLED:
            try:
                service = build("calendar", "v3", credentials=creds)
                tomorrow = (datetime.utcnow() + timedelta(days=1)).date()
                slots = get_available_slots(service, tomorrow, company_tz, lead_tz)
            except Exception as e:
                print(f"Slot computation failed: {e}")
                return jsonify({
                    "error": f"Calendar availability lookup failed ({e}); no call was placed. "
                             "Retry when the calendar can be checked, or book manually.",
                    "status": "availability_failed",
                }), 502
            if not slots:
                return jsonify({
                    "error": "No calendar-confirmed free slots are available tomorrow; no call was placed. "
                             "Offer the lead an alternative day or book manually.",
                    "status": "no_slots",
                }), 502
        # Preview mode (live disabled) skips the lookup entirely: the result is
        # simulated and nothing is dialed or booked.

        # Step 1: Place outbound CALL-E call with structured result extraction
        call_result = place_call_e_call(
            to_number=phone,
            message=generate_call_message(name, your_company, company, lead_tz, slots),
            result_schema=CALL_E_RESULT_SCHEMA,
            idempotency_key=lead_idempotency_key(name, email, phone, company, your_company, "tomorrow"),
        )

        call_sid = call_result.get("sid")

        # Remember the lead so the call-status flow can schedule at the right time
        if call_sid and call_sid != "error":
            pending_leads[call_sid] = {
                "name": name,
                "email": email,
                "phone": phone,
                "company": company,
                "your_company": your_company,
                "company_tz": company_tz,
                "lead_tz": lead_tz,
                "session_id": current_session_id(),
                "simulated": bool(call_result.get("simulated")),
                "consent": consent,
                "consent_recorded_at": datetime.utcnow().isoformat() + "Z",
            }

        # Combine results
        response = {
            "status": "success",
            "call_sid": call_sid,
            "call_status": call_result.get("status"),
            "mode": call_result.get("mode", "simulated"),
            "message": f"Follow-up call initiated for {name} from {your_company}"
        }
        if call_result.get("simulated"):
            response["message"] = (
                f"Simulated follow-up for {name} from {your_company} "
                f"(live calls disabled server-side) - no call was placed and nothing will be booked."
            )

        return jsonify(response)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def booking_decision(status, structured, simulated=False):
    """Fail-closed decision whether a booking may be made.

    Returns (preferred_day, preferred_time, confirmed_tz) or None. A booking
    is possible only when ALL of the following hold:
      - the provider reported the authoritative terminal status 'completed';
      - the call was a real call (never from a simulation/shim);
      - the structured result is a non-empty dict (confirmation evidence);
      - wants_appointment == 'yes' and time_confirmed == 'yes' (explicit);
      - preferred_day is today/tomorrow and preferred_time parses as HH:MM;
      - the lead-confirmed timezone is present and not 'unknown'.
    Anything else returns None and no event is created.
    """
    if simulated:
        return None
    if status != "completed":
        return None
    if not isinstance(structured, dict) or not structured:
        return None
    if structured.get("wants_appointment") != "yes":
        return None
    if structured.get("time_confirmed") != "yes":
        return None
    day = structured.get("preferred_day")
    time_str = structured.get("preferred_time")
    tz = structured.get("timezone")
    if day not in ("today", "tomorrow"):
        return None
    if not time_str or time_str == "unknown" or parse_preferred_time(time_str) is None:
        return None
    if not tz or tz == "unknown":
        return None
    return day, time_str, tz


@app.route("/call-status/<call_id>")
def call_status(call_id):
    """Poll the CALL-E call and, when done, schedule the appointment at the lead's preferred time."""
    try:
        auth_err = require_token(strict=True)
        if auth_err:
            return auth_err

        # Preview mode has no real SID: return an honest simulated result.
        if not call_id or call_id in ("null", "undefined"):
            return jsonify({
                "call_id": call_id,
                "status": "simulated",
                "summary": "Preview mode: no real call was placed.",
                "structured_result": {},
                "calendar": {
                    "status": "simulated",
                    "message": "Nothing was booked in preview mode.",
                },
            })

        lead = pending_leads.get(call_id, {})

        if call_e_client.using_real_sdk:
            raw = call_e_client._impl.calls.get(call_id)
            status = raw.get("status")
            structured = raw.get("structured_result") or {}
            summary = raw.get("summary")
        else:
            status = "simulated"
            structured = {}
            summary = "simulated call"

        response = {
            "call_id": call_id,
            "status": status,
            "summary": summary,
            "structured_result": structured,
        }

        # Only schedule once the call reached a terminal state
        terminal = status in ("completed", "failed", "canceled")
        if terminal and call_id in pending_leads:
            decision = booking_decision(
                status, structured, simulated=bool(lead.get("simulated"))
            )
            if decision is not None:
                preferred_day, preferred_time, confirmed_tz = decision
                response["calendar"] = schedule_google_calendar(
                    lead.get("name", "Valued Customer"),
                    lead.get("email", ""),
                    lead.get("phone", ""),
                    lead.get("company", ""),
                    your_company=lead.get("your_company", "Company"),
                    preferred_day=preferred_day,
                    preferred_time=preferred_time,
                    call_sid=call_id,
                    lead_tz=confirmed_tz,
                    company_tz=lead.get("company_tz", "UTC"),
                )
                # Send confirmation email once the event is on the calendar
                if response["calendar"].get("status") == "success":
                    response["email"] = send_confirmation_email(
                        lead.get("name", "Valued Customer"),
                        lead.get("email", ""),
                        lead.get("company", ""),
                        response["calendar"],
                        lead.get("phone", ""),
                        your_company=lead.get("your_company", "Company"),
                    )
            elif structured and structured.get("time_confirmed") == "no":
                response["calendar"] = {
                    "status": "not_confirmed",
                    "message": "Lead did not confirm the time/timezone - no appointment scheduled.",
                }
            elif structured and structured.get("wants_appointment") != "yes":
                response["calendar"] = {
                    "status": "no_appointment",
                    "message": f"Lead did not request an appointment (wants_appointment={structured.get('wants_appointment')}).",
                }
            else:
                response["calendar"] = {
                    "status": "evidence_required",
                    "message": "Booking skipped: the call result did not meet the fail-closed "
                               "requirements (authoritative completion, explicit time confirmation, "
                               "bound day/time, and confirmed timezone).",
                }
            # Stop tracking once processed
            del pending_leads[call_id]

        return jsonify(response)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
# BATCH UPLOAD (Excel/CSV) - sequential calls + status tracking
# ============================================================

BATCH_HEADERS = ["name", "phone", "email", "company", "your_company", "company_tz", "consent"]
BATCH_ALIASES = {
    "name": ["name", "lead name", "full name", "lead name"],
    "phone": ["phone", "phone number", "mobile", "phone no", "phonenumber"],
    "email": ["email", "email address", "mail"],
    "company": ["company", "lead company", "lead's company", "company name"],
    "your_company": ["your_company", "your company", "caller company", "company a", "your company (caller)"],
    "status": ["status", "result", "outcome", "scheduling status"],
    "company_tz": ["company_tz", "your timezone", "caller timezone", "timezone"],
}


def normalize_header(header):
    """Match an uploaded header to a canonical field name."""
    h = str(header).strip().lower()
    for field, aliases in BATCH_ALIASES.items():
        if h in aliases:
            return field
    return None


def read_batch_rows(file_storage):
    """Parse an uploaded .xlsx/.csv file into a list of row dicts.

    Rows already marked as scheduled/success (from a previous run) are skipped.
    Returns (rows, matched_headers, skipped_count).
    """
    filename = (file_storage.filename or "").lower()
    content = file_storage.read()
    rows = []
    matched = {}
    skipped = 0

    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        lines = list(ws.iter_rows(values_only=True))
    else:  # treat as csv/tsv
        text = content.decode("utf-8-sig", errors="replace")
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t") if text.strip() else None
        reader = csv.reader(io.StringIO(text), dialect or csv.excel)
        lines = list(reader)

    if not lines:
        return [], matched, 0

    # Map header row to canonical fields
    header_row = [normalize_header(c) for c in lines[0]]
    for idx, field in enumerate(header_row):
        if field:
            matched[field] = idx
    status_col = None
    if "status" in matched:
        status_col = matched["status"]

    if "name" not in matched or "phone" not in matched:
        raise ValueError("Uploaded file must have 'name' and 'phone' columns (headers in first row).")

    for line in lines[1:]:
        if not line or all(c is None or str(c).strip() == "" for c in line):
            continue
        row = {}
        for field in BATCH_HEADERS:
            idx = matched.get(field)
            row[field] = str(line[idx]).strip() if idx is not None and idx < len(line) and line[idx] is not None else ""
        # Validate phone to strict E.164; invalid rows are marked 'error'
        # (imported lazily so batch uploads work even without phonenumbers).
        import phonenumbers as _pn  # noqa: F401
        e164 = validate_e164(row["phone"])
        if not e164:
            row["status"] = "error"
            row["error"] = "Invalid phone number (must be a valid E.164 number)."
            row["_invalid"] = True
            rows.append(row)
            continue
        row["phone"] = e164
        row["lead_tz"] = get_lead_timezone(row["phone"])
        row["consent"] = (row.get("consent") or "").strip().lower()

        # Live calls require recorded consent per recipient.
        if LIVE_CALLS_ENABLED and row["consent"] != "yes":
            row["status"] = "error"
            row["error"] = "Consent required: set consent=yes to authorize a live call for this recipient."
            row["_invalid"] = True
            rows.append(row)
            continue
        row["consent_recorded_at"] = datetime.utcnow().isoformat() + "Z"

        # Skip rows already successfully scheduled in a previous run
        if status_col is not None and status_col < len(line) and line[status_col]:
            prev = str(line[status_col]).strip().lower()
            if prev in ("scheduled", "done", "success", "completed"):
                skipped += 1
                continue
        rows.append(row)

    return rows, matched, skipped


def process_batch(batch_id):
    """Background worker: process batch rows sequentially.

    Fail-closed rules:
      - Live mode dials only when the calendar lookup succeeds and at least
        one calendar-confirmed slot is free; lookup failure or an empty slot
        list marks the row 'error' and no call is placed for that recipient.
      - An ambiguous provider outcome (status-fetch error or polling timeout)
        STOPS the batch: no further recipients are called, and the remaining
        rows are marked 'stopped'.
    """
    job = batch_jobs[batch_id]
    stopped = False
    for row in job["rows"]:
        # Rows rejected at upload (e.g. invalid phone) are already terminal.
        if row.get("_invalid") or row.get("status") in ("error", "stopped", "simulated"):
            job["done_count"] += 1
            continue
        if job.get("stop") or stopped:
            row["status"] = "stopped"
            job["done_count"] += 1
            continue

        row["status"] = "calling"
        try:
            # Pre-compute tomorrow's free slots for this lead so the agent
            # can offer only available times. Live mode fails closed before
            # dialing when the lookup fails or nothing is free.
            slots = None
            if LIVE_CALLS_ENABLED:
                creds = get_google_credentials(session_id=job.get("session_id"))
                if not creds:
                    row["status"] = "error"
                    row["error"] = "Google Calendar is not connected; no call placed for this recipient."
                    job["done_count"] += 1
                    continue
                service = build("calendar", "v3", credentials=creds)
                tomorrow = (datetime.utcnow() + timedelta(days=1)).date()
                slots = get_available_slots(
                    service, tomorrow,
                    row.get("company_tz") or "UTC", row.get("lead_tz") or None,
                )
                if not slots:
                    row["status"] = "error"
                    row["error"] = "No calendar-confirmed free slots are available tomorrow; no call placed for this recipient."
                    job["done_count"] += 1
                    continue

            call_result = place_call_e_call(
                to_number=row["phone"],
                message=generate_call_message(row["name"], row["your_company"], row["company"], row.get("lead_tz"), slots),
                result_schema=CALL_E_RESULT_SCHEMA,
                idempotency_key=lead_idempotency_key(
                    row["name"], row["email"], row["phone"], row["company"],
                    row["your_company"], "tomorrow",
                ),
            )
            row["call_sid"] = call_result.get("sid")
            row["simulated"] = bool(call_result.get("simulated"))
            if call_result.get("simulated"):
                row["status"] = "simulated"
                row["error"] = "Live calls disabled server-side (CALLE_LIVE_CALLS_ENABLED) - no call placed."
                job["done_count"] += 1
                continue
            row["status"] = "waiting"

            # Poll until terminal state (up to ~10 min)
            terminal = False
            ambiguous = False
            status = None
            structured = {}
            for _ in range(40):
                time.sleep(15)
                if job.get("stop"):
                    break
                try:
                    if call_e_client.using_real_sdk:
                        raw = call_e_client._impl.calls.get(row["call_sid"])
                        status = raw.get("status")
                        if not status:
                            raise ValueError("provider returned no status")
                        structured = raw.get("structured_result") or {}
                    else:
                        status = "simulated"
                        structured = {}
                    if status in ("completed", "failed", "canceled"):
                        row["call_status"] = status
                        row["structured"] = structured
                        terminal = True
                        break
                except Exception as e:
                    row["status"] = "error"
                    row["error"] = f"Call status lookup failed: {e}"
                    ambiguous = True
                    terminal = True
                    break

            if job.get("stop"):
                row["status"] = "stopped"
            elif ambiguous:
                # An ambiguous provider outcome stops the batch: a call may
                # still be in flight, so no further recipients are dialed.
                job["status"] = "stopped"
                job["stopped_reason"] = (
                    f"Batch stopped: call status lookup failed for {mask_phone(row['phone'])} "
                    f"({row['error']}); no further calls were placed."
                )
                stopped = True
            elif not terminal:
                row["status"] = "error"
                row["error"] = "Call status unknown (polling timed out); no booking attempted."
                job["status"] = "stopped"
                job["stopped_reason"] = (
                    "Batch stopped: polling timed out while awaiting a terminal call status; "
                    "no further calls were placed."
                )
                stopped = True
            elif terminal and status == "completed":
                decision = booking_decision(status, structured, simulated=bool(row.get("simulated")))
                if decision is not None:
                    preferred_day, preferred_time, confirmed_tz = decision
                    cal = schedule_google_calendar(
                        row["name"], row["email"], row["phone"], row["company"],
                        your_company=row["your_company"],
                        preferred_day=preferred_day, preferred_time=preferred_time,
                        call_sid=row["call_sid"],
                        lead_tz=confirmed_tz,
                        company_tz=row.get("company_tz") or "UTC",
                    )
                    if cal.get("status") == "success":
                        row["status"] = "scheduled"
                        row["event_id"] = cal.get("event_id")
                        row["htmlLink"] = cal.get("htmlLink")
                        row["appointment"] = f"{preferred_day} {preferred_time}"
                        email = send_confirmation_email(
                            row["name"], row["email"], row["company"], cal,
                            row["phone"], your_company=row["your_company"],
                        )
                        row["email_status"] = email.get("status", "unknown")
                        if email.get("status") == "error":
                            row["email_error"] = email.get("message")
                    else:
                        row["status"] = "error"
                        row["error"] = cal.get("message", "scheduling failed")
                elif structured and structured.get("time_confirmed") == "no":
                    row["status"] = "not_confirmed"
                    row["error"] = "Lead did not confirm the time/timezone."
                elif structured and structured.get("wants_appointment") != "yes":
                    row["status"] = "declined"
                    row["error"] = "Lead did not request an appointment."
                else:
                    row["status"] = "no_time"
                    row["error"] = "Booking skipped: call result lacked the fail-closed evidence (authoritative completion, explicit time confirmation, bound day/time, confirmed timezone)."
            elif terminal and status == "failed":
                row["status"] = "failed"
                row["error"] = "Call did not connect."
            elif terminal:
                row["status"] = "declined"
        except Exception as e:
            row["status"] = "error"
            row["error"] = str(e)

        job["done_count"] += 1

    job["running"] = False


@app.route("/batch-upload", methods=["POST"])
def batch_upload():
    """Upload an Excel/CSV file and process leads sequentially."""
    try:
        auth_err = require_token(strict=True)
        if auth_err:
            return auth_err

        if LIVE_CALLS_ENABLED:
            confirm = request.headers.get("X-Confirm-Live-Call", "")
            if confirm != "I understand this places a real phone call":
                return jsonify({
                    "error": "Live calls require the confirmation header "
                             "'X-Confirm-Live-Call: I understand this places a real phone call'.",
                }), 403

        if LIVE_CALLS_ENABLED:
            creds = get_google_credentials()
            if not creds:
                return jsonify({"error": "Google not connected. Visit /oauth first."}), 403

        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"error": "No file uploaded"}), 400

        try:
            rows, matched, skipped = read_batch_rows(file)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

        if not rows:
            return jsonify({"error": "No new leads to process (all were already scheduled in a previous run)."}), 200

        batch_id = f"batch_{int(time.time())}_{len(batch_jobs)}"
        batch_jobs[batch_id] = {
            "rows": rows,
            "matched": matched,
            "total": len(rows),
            "done_count": 0,
            "running": True,
            "stop": False,
            "session_id": current_session_id(),
            "created": datetime.utcnow().isoformat(),
        }

        t = threading.Thread(target=process_batch, args=(batch_id,), daemon=True)
        t.start()

        return jsonify({
            "status": "success",
            "batch_id": batch_id,
            "total": len(rows),
            "skipped": skipped,
            "message": f"Batch started with {len(rows)} leads (skipped {skipped} already scheduled).",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/batch-status/<batch_id>")
def batch_status(batch_id):
    """Return current progress of a batch job (phone numbers masked)."""
    auth_err = require_token(strict=True)
    if auth_err:
        return auth_err
    job = batch_jobs.get(batch_id)
    if not job:
        return jsonify({"error": "Batch not found"}), 404

    return jsonify({
        "batch_id": batch_id,
        "total": job["total"],
        "done_count": job["done_count"],
        "running": job["running"],
        "status": job.get("status", "running"),
        "stopped_reason": job.get("stopped_reason"),
        "rows": [
            {
                "name": r["name"],
                "phone": mask_phone(r["phone"]),
                "company": r["company"],
                "your_company": r["your_company"],
                "company_tz": r.get("company_tz") or "UTC",
                "lead_tz": r.get("lead_tz"),
                "consent": r.get("consent", ""),
                "consent_recorded_at": r.get("consent_recorded_at"),
                "status": r.get("status", "pending"),
                "call_sid": r.get("call_sid"),
                "appointment": r.get("appointment"),
                "event_id": r.get("event_id"),
                "htmlLink": r.get("htmlLink"),
                "error": r.get("error"),
            }
            for r in job["rows"]
        ],
    })


@app.route("/batch-stop/<batch_id>", methods=["POST"])
def batch_stop(batch_id):
    """Stop a running batch after the current call completes."""
    auth_err = require_token(strict=True)
    if auth_err:
        return auth_err
    job = batch_jobs.get(batch_id)
    if not job:
        return jsonify({"error": "Batch not found"}), 404
    job["stop"] = True
    return jsonify({"status": "stopping"})


@app.route("/batch-download/<batch_id>")
def batch_download(batch_id):
    """Download the batch as an updated Excel file with a status column.

    Phone numbers are masked in the exported file for privacy.
    """
    auth_err = require_token(strict=True)
    if auth_err:
        return auth_err
    job = batch_jobs.get(batch_id)
    if not job:
        return jsonify({"error": "Batch not found"}), 404

    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = BATCH_HEADERS + ["consent_recorded_at", "status", "appointment", "call_sid", "event_id", "calendar_link", "error"]
    ws.append(headers)

    for r in job["rows"]:
        ws.append([
            r.get("name", ""), mask_phone(r.get("phone", "")), r.get("email", ""),
            r.get("company", ""), r.get("your_company", ""), r.get("company_tz", ""),
            r.get("consent", ""), r.get("consent_recorded_at", ""),
            r.get("status", "pending"), r.get("appointment", ""),
            r.get("call_sid", ""), r.get("event_id", ""),
            r.get("htmlLink", ""), r.get("error", ""),
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"batch_{batch_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/oauth2callback")
def oauth2callback():
    """Handle Google OAuth callback.

    This route is reached by browser navigation, so bearer auth cannot be
    attached. It is authenticated by the state parameter: the callback is
    accepted only when the state matches the one stored in this browser's
    signed session cookie, and the exchange uses PKCE. Everything else fails
    closed with a 400.
    """
    auth_code = request.args.get("code")
    state = request.args.get("state", "")
    error = request.args.get("error", "")

    if error:
        return jsonify({"error": f"OAuth error: {error}"}), 400
    if not auth_code:
        return jsonify({"error": "No authorization code provided"}), 400

    expected_state = session.pop("oauth_state", None)
    code_verifier = session.pop("oauth_code_verifier", None)
    if not expected_state or not code_verifier:
        return jsonify({"error": "OAuth state missing - start the flow from the app."}), 400
    if state != expected_state:
        return jsonify({"error": "OAuth state mismatch - the callback was rejected."}), 400

    try:
        from google_auth_oauthlib.flow import Flow

        flow = Flow.from_client_config(
            GOOGLE_OAUTH_CONFIG,
            scopes=GOOGLE_SCOPES,
            autogenerate_code_verifier=False,
        )
        flow.redirect_uri = GOOGLE_REDIRECT_URI
        flow.code_verifier = code_verifier

        flow.fetch_token(code=auth_code)
        creds = flow.credentials

        # Save per-session; disk persistence only on explicit operator opt-in.
        save_credentials(creds)
        if os.path.exists(DISCONNECT_FLAG):
            os.remove(DISCONNECT_FLAG)

        return redirect("/?oauth=connected")

    except Exception as e:
        return redirect("/?oauth=error&message=" + str(e))


@app.route("/detect-timezone")
def detect_timezone():
    """Return the timezone derived from a valid E.164 phone number."""
    auth_err = require_token(strict=True)
    if auth_err:
        return auth_err
    phone = request.args.get("phone", "")
    if not phone:
        return jsonify({"timezone": None, "label": None})
    if not validate_e164(phone):
        return jsonify({"error": "Invalid phone number (must be a valid E.164 number)."}), 400
    try:
        tz = get_lead_timezone(phone)
        return jsonify({"timezone": tz, "label": tz_friendly_name(tz)})
    except Exception:
        return jsonify({"timezone": None, "label": None})


@app.route("/oauth-status")
def oauth_status():
    """Return whether Google is connected and with which account/scopes (account masked)."""
    auth_err = require_token(strict=True)
    if auth_err:
        return auth_err

    creds = get_google_credentials()
    if not creds:
        return jsonify({"connected": False, "account": None, "scopes": []})

    scopes = list(getattr(creds, "scopes", []) or [])
    account = None
    try:
        service = build("gmail", "v1", credentials=creds)
        profile = service.users().getProfile(userId="me").execute()
        email = profile.get("emailAddress") or ""
        if "@" in email:
            local, _, domain = email.partition("@")
            account = local[:2] + "***@" + domain
        else:
            account = None
    except Exception:
        pass

    return jsonify({
        "connected": True,
        "account": account,
        "scopes": scopes,
    })


@app.route("/oauth-disconnect", methods=["POST"])
def oauth_disconnect():
    """Revoke the calling session's Google credentials so the user can reconnect."""
    auth_err = require_token(strict=True)
    if auth_err:
        return auth_err
    try:
        sid = current_session_id()
        oauth_tokens.pop(sid, None)
        if PERSIST_TOKEN and os.path.exists(CREDENTIALS_FILE):
            os.remove(CREDENTIALS_FILE)
        with open(DISCONNECT_FLAG, "w") as f:
            f.write("disconnected")
        return jsonify({"status": "success", "connected": False})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/oauth")
def oauth_start():
    """Start Google OAuth flow with state validation and PKCE."""
    auth_err = require_token(strict=True)
    if auth_err:
        return auth_err
    try:
        from google_auth_oauthlib.flow import Flow

        flow = Flow.from_client_config(
            GOOGLE_OAUTH_CONFIG,
            scopes=GOOGLE_SCOPES,
            autogenerate_code_verifier=True,
        )
        flow.redirect_uri = GOOGLE_REDIRECT_URI

        state = os.urandom(16).hex()
        session["oauth_state"] = state
        session["oauth_code_verifier"] = flow.code_verifier

        auth_url, _ = flow.authorization_url(
            access_type="offline",
            prompt="consent",
            state=state,
        )
        return jsonify({"url": auth_url})
    except Exception as e:
        return jsonify({"error": f"OAuth start failed: {str(e)}"}), 500


# ============================================================
# UTILITY: OAuth Setup Script
# ============================================================

def run_oauth_setup():
    """Run OAuth flow to get Google credentials (PKCE, state validated on completion)."""
    import sys
    
    flow = Flow.from_client_config(
        GOOGLE_OAUTH_CONFIG,
        scopes=GOOGLE_SCOPES,
        autogenerate_code_verifier=True,
    )
    flow.redirect_uri = GOOGLE_REDIRECT_URI

    # Generate authorization URL with a state token validated on callback
    state = os.urandom(16).hex()
    auth_url, _ = flow.authorization_url(access_type="offline", prompt="consent", state=state)
    
    print("=" * 60)
    print("GOOGLE OAUTH SETUP FOR CALL-E HACKATHON")
    print("=" * 60)
    print(f"\n1. Open this URL in your browser: {auth_url}")
    print("2. Log in with your Google account and authorize access")
    print("3. You'll be redirected to: http://localhost:8080/oauth2callback")
    print("4. Copy the 'code' parameter from the URL")
    print("5. Run: python app.py oauth2callback?code=YOUR_CODE_HERE")
    print("=" * 60)


# ============================================================
# MAIN ENTRY POINT
# ============================================================

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    
    # Check if we need to run OAuth setup
    if len(sys.argv) > 1 and sys.argv[1] == "oauth":
        run_oauth_setup()
        sys.exit(0)
    
    print("=" * 60)
    print("CALL-E Hackathon: Customer Follow-Up Automation")
    print("Hybrid: Python SDK + Google Calendar API")
    print("=" * 60)
    if not LIVE_CALLS_ENABLED:
        print("PREVIEW MODE: live calls are DISABLED (CALLE_LIVE_CALLS_ENABLED unset).")
        print("All call results are simulated; nothing is dialed or booked.")
        if not APP_TOKEN:
            print("No APP_TOKEN configured: only lead submission runs unauthenticated.")
            print("Call-status, OAuth, batch, and data routes still require authentication.")
        else:
            print("APP_TOKEN is configured and enforced on all private routes.")
    else:
        print("LIVE MODE: CALLE_LIVE_CALLS_ENABLED=true. Real calls can be placed.")
        print("Every live request requires the X-Confirm-Live-Call header and a Bearer APP_TOKEN.")
    print(f"Server starting at http://localhost:{port}")
    print("Endpoints:")
    print("  GET  /          - Demo information page")
    print("  POST /lead-submission - Submit lead data (triggers CALL-E call)")
    print("  GET /oauth2callback - Google OAuth completion")
    print("  GET app.py oauth - OAuth setup guide")
    print("=" * 60)
    print("\nDemo Flow:")
    print("1. Complete OAuth: python app.py oauth")
    print("2. GET / to see setup status")
    print("3. POST lead data to /lead-submission")
    print("4. CALL-E SDK places outbound call automatically")
    print("5. Personalized message plays with scheduling option")
    print("6. Press 1 -> Google Calendar event created")
    print("7. Confirmation email/SMS sent")
    print("=" * 60)
    app.run(host="0.0.0.0", port=port, debug=False)